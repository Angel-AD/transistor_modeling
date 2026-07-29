"""
keep_columns.py  --  Write a slimmed copy of an xlsx keeping only specified columns.

Reads the source .xlsx (header in row 1) and writes a NEW <name><suffix>.xlsx containing
only the --keep columns, in the order given. The source is never modified. Columns in
--keep that aren't present in a file are skipped (with a note), so it's safe across files
with slightly different schemas.

Usage
-----
    python keep_columns.py --xlsx a.xlsx b.xlsx
    python keep_columns.py --xlsx a.xlsx --keep id eq combined_gm ids_rmse --suffix _slim
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# Default columns to keep (order preserved in the output).
DEFAULT_KEEP = ["id", "eq", "combined_gm", "ids_rmse",
                "region_knee_combined_gm", "region_knee_ids_rmse",
                "region_weight", "region_vds_hi", "region_vds_lo",
                "region_vgs_hi", "region_vgs_lo", "gm_vds_min", "gm_vgs_min"]

_FILL = PatternFill("solid", fgColor="1F4E78")
_FONT = Font(bold=True, color="FFFFFF")
_ALIGN = Alignment(horizontal="center", vertical="center")


def slim_xlsx(src: Path, keep, suffix: str):
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        print(f"  [skip] {src.name}: empty")
        return None
    header = [("" if h is None else str(h)) for h in rows[0]]
    idx = [header.index(c) for c in keep if c in header]
    kept = [header[i] for i in idx]
    missing = [c for c in keep if c not in header]

    out = src.with_name(src.stem + suffix + ".xlsx")
    ow = openpyxl.Workbook()
    os_ = ow.active
    os_.title = "slim"
    os_.append(kept)
    for r in rows[1:]:
        os_.append([(r[i] if i < len(r) else None) for i in idx])
    for c in range(1, len(kept) + 1):
        cell = os_.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment = _FILL, _FONT, _ALIGN
    os_.freeze_panes = "A2"
    try:
        ow.save(out)
    except PermissionError:
        print(f"  [skip] {out.name} is open/locked (e.g. in Excel) -- close it and re-run.")
        return None
    print(f"  {src.name} -> {out.name}  ({len(kept)} cols, {len(rows) - 1} rows"
          + (f"; missing: {missing}" if missing else "") + ")")
    return out


def main():
    ap = argparse.ArgumentParser(
        description="Write slimmed xlsx copies keeping only specified columns.",
        formatter_class=argparse.RawDescriptionHelpFormatter, epilog=__doc__)
    ap.add_argument("--xlsx", nargs="+", required=True, help="Source xlsx file(s).")
    ap.add_argument("--keep", nargs="+", default=DEFAULT_KEEP,
                    help="Columns to keep, in order. Default: the region-analysis set.")
    ap.add_argument("--suffix", default="_slim", help="Output filename suffix (default _slim).")
    args = ap.parse_args()

    print(f"keep: {args.keep}")
    n = 0
    for x in args.xlsx:
        p = Path(x)
        if not p.is_file():
            print(f"  [skip] not found: {p}")
            continue
        if slim_xlsx(p, args.keep, args.suffix):
            n += 1
    print(f"\ndone. wrote {n} slim xlsx.")


if __name__ == "__main__":
    main()
