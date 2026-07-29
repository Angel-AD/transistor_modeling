"""
Render compiled-results CSVs as formatted Excel workbooks.

Reads CSVs produced by compile_results_csv.py and/or compile_master_best.py
and writes a sibling .xlsx with:
  - bold white-on-dark-blue header row, centered, with thin border;
  - scientific notation (0.000E+00) for loss / RMSE / lr columns;
  - frozen header row;
  - auto-sized column widths;
  - a thick top border between experiment groups (only for files with an
    'experiment' column, i.e. master_best_*.csv; other files are unaffected).

Usage:
    # Convert a single CSV
    python compile_results_to_xlsx.py --csv path/to/compiled_results.csv

    # Convert every compiled_results.csv + master_best_*.csv under a root
    python compile_results_to_xlsx.py --root path/to/experiment_root

Requires `openpyxl` (pip install openpyxl).
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl",
          file=sys.stderr)
    raise

# Columns that should be rendered in scientific notation when populated.
SCIENTIFIC_COLUMNS = {
    "best_loss", "mse_loss", "ids_rmse",
    "gm1_rmse", "gm2_rmse", "gm3_rmse",
    "lr", "learning_rate",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")  # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
# Thick top border to visually separate experiment groups (master_best_*.csv,
# which have an 'experiment' column with contiguous per-experiment blocks).
THICK_TOP_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thick", color="000000"),
    bottom=Side(style="thin", color="BFBFBF"),
)
SCI_FORMAT = "0.000E+00"


def _coerce(value: str):
    """Best-effort CSV string -> int/float/str conversion. Empty stays empty."""
    if value is None or value == "":
        return None
    try:
        if "." in value or "e" in value or "E" in value or "inf" in value.lower():
            return float(value)
        return int(value)
    except ValueError:
        return value


def csv_to_xlsx(csv_path: Path, xlsx_path: Path | None = None) -> Path:
    if xlsx_path is None:
        xlsx_path = csv_path.with_suffix(".xlsx")

    with csv_path.open("r", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        raise ValueError(f"{csv_path} is empty")

    header, data = rows[0], rows[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = csv_path.stem[:31] or "results"

    # --- Header row ---
    for ci, name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    sci_cols = {i for i, h in enumerate(header, start=1) if h in SCIENTIFIC_COLUMNS}

    # If rows are grouped by experiment (compile_master_best output has an
    # 'experiment' column with contiguous per-experiment blocks), draw a thick top
    # border on the first row of each new group to separate experiments visually.
    # Files without an 'experiment' column (ranked_* / single-experiment) are unaffected.
    try:
        exp_idx = header.index("experiment")
    except ValueError:
        exp_idx = None
    prev_exp = None

    # --- Data rows ---
    for ri, row in enumerate(data, start=2):
        # Pad short rows so column iteration is safe.
        if len(row) < len(header):
            row = row + [""] * (len(header) - len(row))
        cur_exp = row[exp_idx] if exp_idx is not None else None
        new_group = exp_idx is not None and ri > 2 and cur_exp != prev_exp
        row_border = THICK_TOP_BORDER if new_group else THIN_BORDER
        for ci, value in enumerate(row[: len(header)], start=1):
            cell = ws.cell(row=ri, column=ci, value=_coerce(value))
            cell.border = row_border
            if ci in sci_cols and isinstance(cell.value, (int, float)):
                cell.number_format = SCI_FORMAT
        prev_exp = cur_exp

    # --- Column widths (cap to keep wide file paths readable but bounded) ---
    for ci, name in enumerate(header, start=1):
        max_len = len(name)
        for row in data:
            if ci - 1 < len(row):
                max_len = max(max_len, len(str(row[ci - 1])))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 10), 60)

    ws.freeze_panes = "A2"

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return xlsx_path


def collect_csvs_under_root(root: Path) -> list[Path]:
    """Find compiled_results.csv (one per experiment subdir) + master_best_*.csv
    (sitting in the root). Returns absolute paths."""
    if not root.is_dir():
        raise SystemExit(f"Root path is not a directory: {root}")
    found: list[Path] = []
    for p in sorted(root.iterdir()):
        if p.is_dir():
            # compile_results_csv.py --root writes compiled_results_<exp>.csv;
            # --dir writes compiled_results.csv. Match both.
            found.extend(sorted(p.glob("compiled_results*.csv")))
    found.extend(sorted(root.glob("master_best_*.csv")))
    found.extend(sorted(root.glob("*_ranked_by_*.csv")))              # legacy: root-level
    # compile_ranked.py groups ranked CSVs into ranked_global/ and ranked_region_<name>/.
    found.extend(sorted(root.glob("ranked_*/*_ranked_by_*.csv")))
    return found


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    src = parser.add_mutually_exclusive_group(required=True)
    src.add_argument("--csv", type=str, help="Path to a single CSV to convert.")
    src.add_argument("--root", type=str,
                     help="Root containing experiment subdirs (each with "
                          "compiled_results.csv) and/or master_best_*.csv.")
    parser.add_argument("--output", type=str, default=None,
                        help="(With --csv only) explicit output xlsx path. "
                             "Defaults to <csv>.xlsx alongside the input.")
    args = parser.parse_args()

    targets: list[tuple[Path, Path | None]] = []
    if args.csv:
        csv_path = Path(args.csv).resolve()
        if not csv_path.is_file():
            raise SystemExit(f"CSV not found: {csv_path}")
        out = Path(args.output).resolve() if args.output else None
        targets.append((csv_path, out))
    else:
        for p in collect_csvs_under_root(Path(args.root).resolve()):
            targets.append((p, None))
        if not targets:
            print(f"No CSVs to convert under {args.root}")
            return 0

    for csv_path, out in targets:
        xlsx = csv_to_xlsx(csv_path, out)
        print(f"  wrote {xlsx}")

    print(f"\nConverted {len(targets)} CSV(s).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
