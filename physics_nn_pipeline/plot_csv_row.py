"""
Plot config(s) selected from a compiled / ranked CSV or xlsx row.

Companion to plot_best_configs.py. Where plot_best plots the top-N by a metric, this
picks explicit rows -- by their `id`, by physical row number, by config_name, or (with no
selector) EVERY row -- and regenerates each run's 6-panel IV/GM plot via plot_saved_state.py.

The `id` column (added by compile_ranked.py / compile_results_csv.py) is a simple 1-based
number written into the file (1 = best by that file's metric). Because it's a stored
column value, it SURVIVES re-sorting the CSV in Excel: sort/reorder the rows however you
like, copy the `id` of the row you want, and pass it with the same --ranked_csv.

Each row's `file_path` column points at the run's run_loss_*.json; its parent directory
is what plot_saved_state.py needs as --dir.

Usage:
    # by id (the 'id' column -- survives Excel re-sorting)
    python plot_csv_row.py --ranked_csv refine2/refine2_ranked_by_ids_rmse.csv \
        --csv ../csvs/<measurements>.csv --id 1

    # several at once
    python plot_csv_row.py --ranked_csv ... --csv ... --id 1,2,3

    # by current physical row position (1-based)
    python plot_csv_row.py --ranked_csv ... --csv ... --row 5

    # by config_name
    python plot_csv_row.py --ranked_csv ... --csv ... --config_name PhysNN_classic_W1.0-1.0-1.0_none

    # every row in the file (csv OR xlsx) -- no selector given
    python plot_csv_row.py --ranked_csv some_slim.xlsx --csv ...

    # every row except a few
    python plot_csv_row.py --ranked_csv some_slim.xlsx --csv ... --skip_ids 3,7,12
"""
from __future__ import annotations

import argparse
import csv
import os
import shutil
import subprocess
import sys
import time
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
PLOT_SCRIPT = HERE / "plot_saved_state.py"

# Base roots to search when a stored file_path no longer exists — e.g. the run
# dirs were moved out of the scripts folder into a sibling runs/ tree.
_REROOT_BASES = (HERE, HERE.parent, HERE.parent / "runs")


def _reroot(p: Path) -> Path:
    """If p is missing, return the longest tail of p that exists under a known
    base root; otherwise return p unchanged."""
    if p.exists():
        return p
    parts = p.parts
    for base in _REROOT_BASES:
        for i in range(len(parts)):
            cand = base.joinpath(*parts[i:])
            if cand.exists():
                return cand
    return p


def _run_dir_from_row(row: dict):
    """Per-run directory from a row's file_path column (parent of run_loss_*.json)."""
    fp = (row.get("file_path") or "").strip()
    if not fp:
        return None
    parent = _reroot(Path(fp).parent)
    return parent if parent.is_dir() else None


# analyze_shape.py's write_gmshape_outputs() derived-file suffixes (see its docstring) --
# each of these sits alongside a "<base>_shape.csv" in the same folder that has the SAME
# ids with a file_path column (the derived files are slim-xlsx-only by convention and
# intentionally drop file_path to stay compact).
_SHAPE_DERIVED_SUFFIXES = (
    "_gmshape_ok_by_gmshape", "_gmshape_ok_by_gdshape",
    "_gdsshape_ok_by_gmshape", "_gdsshape_ok_by_gdshape",
    "_bothshape_ok", "_bothshape_removed",
)


def _find_base_shape_csv(path: Path) -> Path | None:
    """If `path` looks like one of analyze_shape.py's derived shape files (e.g.
    '<base>_shape_gmshape_ok_by_gmshape_slim.xlsx'), return the sibling '<base>_shape.csv'
    that has the same ids plus a file_path column -- else None."""
    stem = path.stem.removesuffix("_slim")
    for suf in _SHAPE_DERIVED_SUFFIXES:
        if stem.endswith(suf):
            base = path.with_name(stem[: -len(suf)] + ".csv")
            return base if base.is_file() else None
    return None


def _read_rows(path: Path) -> list[dict]:
    """List of header-keyed dicts from a .csv or .xlsx ranked/compiled/slim file."""
    if path.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [("" if h is None else str(h)) for h in next(it)]
        rows = [dict(zip(header, r)) for r in it]
        wb.close()
        return rows
    with open(path, "r", newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _select(rows: list[dict], args) -> list[dict]:
    if args.id:
        if "id" not in rows[0]:
            raise SystemExit("This CSV has no 'id' column -- regenerate it with the "
                             "updated compile_ranked.py / compile_results_csv.py.")
        by = {str(r["id"]): r for r in rows}
        out = []
        for u in [x.strip() for x in args.id.split(",") if x.strip()]:
            if u in by:
                out.append(by[u])
            else:
                print(f"[warn] id {u!r} not found in CSV")
    elif args.run_id:
        if "run_id" not in rows[0]:
            raise SystemExit("This CSV has no 'run_id' column -- regenerate it with the "
                             "updated compile_ranked.py / compile_results_csv.py.")
        by = {str(r["run_id"]): r for r in rows}
        out = []
        for u in [x.strip() for x in args.run_id.split(",") if x.strip()]:
            if u in by:
                out.append(by[u])
            else:
                print(f"[warn] run_id {u!r} not found in CSV")
    elif args.row:
        out = []
        for x in [x.strip() for x in args.row.split(",") if x.strip()]:
            i = int(x)
            if 1 <= i <= len(rows):
                out.append(rows[i - 1])
            else:
                print(f"[warn] row {i} out of range (1..{len(rows)})")
    elif args.config_name:
        want = {c.strip() for c in args.config_name.split(",") if c.strip()}
        out = [r for r in rows if (r.get("config_name") or "") in want]
    else:
        out = list(rows)   # no selector given -> every row

    if args.skip_ids:
        if "id" not in rows[0]:
            raise SystemExit("--skip_ids needs an 'id' column, which this CSV doesn't have.")
        skip = {x.strip() for x in args.skip_ids.split(",") if x.strip()}
        before = len(out)
        out = [r for r in out if str(r.get("id")) not in skip]
        print(f"[skip_ids] excluded {before - len(out)}/{before} row(s): {sorted(skip)}")
    return out


def main():
    ap = argparse.ArgumentParser(
        description="Plot specific config(s) chosen from a CSV row (by uid / row / config_name).",
        formatter_class=argparse.RawDescriptionHelpFormatter, epilog=__doc__)
    ap.add_argument("--ranked_csv", required=True,
                    help="CSV or xlsx holding the rows (ranked_by_*.csv/.xlsx, compiled_results*.csv, "
                         "master_best_*.csv, or a keep_columns.py _slim.xlsx).")
    ap.add_argument("--csv", default=None,
                    help="Measurement CSV (forwarded to plot_saved_state.py). If omitted, "
                         "auto-detected from the ranked CSV's results folder "
                         "(base_files/_run_meta.json).")
    sel = ap.add_mutually_exclusive_group(required=False)
    sel.add_argument("--id", help="Row id(s) from the 'id' column, comma-separated. Survives Excel re-sorting.")
    sel.add_argument("--run_id", help="Row id(s) from the GLOBAL 'run_id' column, comma-separated. "
                     "Unlike 'id' (which renumbers per ranking/sort order), run_id is stable across "
                     "different ranked files -- and across different results folders that share the "
                     "same config (same architecture -> same exp_NNN index -> same run_id), so it's "
                     "the right selector for comparing one architecture's fit across multiple csvs.")
    sel.add_argument("--row", help="Current physical row number(s), 1-based, comma-separated.")
    sel.add_argument("--config_name", help="Match by config_name (comma-separated).")
    ap.add_argument("--skip_ids", default=None,
                    help="Comma-separated 'id' values to exclude from the selected set. Most useful "
                         "with no selector above (plots EVERY row in the file) to skip a few; also "
                         "applies on top of --id/--run_id/--row/--config_name.")
    ap.add_argument("--min_vgs", type=float, default=-4.0)
    ap.add_argument("--plot_vds_list", type=str, default="0,5,10,15,20,28")
    ap.add_argument("--plot_vgs_list", type=str, default="-3.5,-3,-2.9,-2.8,-2.7,-2.5,-2.2,-2,-1.8,-1.6,-1.5,-1.3,-1,-.5,0")
    ap.add_argument("--val", type=str, default=None,
                    help="Validation set option forwarded to plot_saved_state.py: "
                         "'interpolation', a float (e.g. '0.3'), or a path to a val CSV.")
    ap.add_argument("--add_zero_vds", action="store_true",
                    help="Forward --add_zero_vds to plot_saved_state.py.")
    ap.add_argument("--python_exe", type=str, default=sys.executable)
    ap.add_argument("--open", action="store_true",
                    help="Open plot_saved_state_full.png in the default image viewer. Smart: if the "
                         "image already exists it just shows it (no re-plot); else it plots then opens. "
                         "Use --force to always re-plot.")
    ap.add_argument("--force", action="store_true",
                    help="With --open, re-plot even if the image already exists.")
    ap.add_argument("--show_only", action="store_true",
                    help="Never plot -- only open the already-created plot_saved_state_full.png for "
                         "each selected id. No measurement CSV needed.")
    ap.add_argument("--dry_run", action="store_true", help="Print the commands, then exit.")
    args = ap.parse_args()

    csv_path = Path(args.ranked_csv).resolve()
    if not csv_path.is_file():
        raise SystemExit(f"--ranked_csv not found: {csv_path}")
    if not PLOT_SCRIPT.is_file():
        raise SystemExit(f"plot_saved_state.py not found: {PLOT_SCRIPT}")

    if not args.csv and not args.show_only:   # auto-detect the measurement CSV from the results folder
        sys.path.insert(0, str(HERE))
        import run_artifacts
        # ranked CSVs live in <root>/ranked_*/f.csv; master_best/compiled in <root>/f.csv
        for root in (csv_path.parent.parent, csv_path.parent):
            det = run_artifacts.run_meta(str(root)).get("csv")
            if det:
                args.csv = det
                print(f"[auto] measurement CSV: {det}")
                break
    if not args.csv and not args.show_only:
        raise SystemExit("--csv not given and could not be auto-detected from the ranked "
                         "CSV's folder (no base_files/_run_meta.json or run records found).")

    rows = _read_rows(csv_path)
    if not rows:
        raise SystemExit(f"No rows in {csv_path}")

    selected = _select(rows, args)
    if not selected:
        raise SystemExit("No matching rows selected.")
    if not args.show_only and "file_path" not in rows[0]:
        base_csv = _find_base_shape_csv(csv_path)
        if base_csv is None:
            guess = csv_path.with_name(csv_path.stem.removesuffix("_slim") + ".csv")
            raise SystemExit(f"{csv_path.name} has no 'file_path' column -- it's likely a _slim "
                             "xlsx (slim files intentionally drop file_path to stay compact), "
                             "and no sibling '<base>_shape.csv' was found to recover it from. "
                             f"Use the corresponding full .csv instead, e.g.: {guess.name}")
        print(f"[auto] {csv_path.name} has no 'file_path' column -- recovering it by id from "
              f"{base_csv.name} (the base shape file these compliance rows were derived from).")
        base_rows = _read_rows(base_csv)
        if "id" not in base_rows[0] or "file_path" not in base_rows[0]:
            raise SystemExit(f"{base_csv.name} is missing 'id' or 'file_path' -- can't recover.")
        fp_by_id = {str(r["id"]): r["file_path"] for r in base_rows}
        missing = []
        for r in selected:
            fp = fp_by_id.get(str(r.get("id")))
            if fp is None:
                missing.append(r.get("id"))
            else:
                r["file_path"] = fp
        if missing:
            raise SystemExit(f"id(s) {missing} not found in {base_csv.name} -- can't recover file_path.")

    ok = fail = 0
    for r in selected:
        label = f"id={r.get('id')}" if r.get("id") else (r.get("config_name") or "?")
        id_val = str(r.get('id', r.get('config_name', label)))
        out_dir = csv_path.parent / "plotted_configs" / f"{csv_path.stem}_id{id_val}"
        cached = out_dir / "plot_saved_state_full.png"

        # Reuse the already-created image instead of re-plotting when:
        #   --show_only  (never plot), or  --open and it already exists (unless --force).
        if args.show_only or (args.open and cached.is_file() and not args.force):
            if cached.is_file():
                print(f"[{label}] showing (cached) {cached}", flush=True)
                if not args.dry_run:
                    try:
                        os.startfile(str(cached))
                        time.sleep(0.7)   # let the viewer open a window before the next request
                                          # (Windows Photos is single-instance; rapid os.startfile
                                          # calls otherwise just redirect one window, so only the
                                          # last image ends up visible for a multi-id --open/--show_only)
                    except Exception as e:
                        print(f"[{label}] could not open {cached}: {e}", flush=True)
                ok += 1
            else:   # only reachable via --show_only
                print(f"[{label}] no existing image at {cached}  "
                      f"(run without --show_only to create it)", flush=True)
                fail += 1
            continue

        run_dir = _run_dir_from_row(r)
        if run_dir is None:
            print(f"[{label}] SKIP: cannot resolve run dir from file_path={r.get('file_path','')!r}")
            fail += 1
            continue
        cmd = [args.python_exe, str(PLOT_SCRIPT), "--dir", str(run_dir), "--csv", args.csv,
               "--min_vgs", str(args.min_vgs),
               f"--plot_vds_list={args.plot_vds_list}", f"--plot_vgs_list={args.plot_vgs_list}"]
        if args.val:
            cmd += ["--val", args.val]
        if args.add_zero_vds:
            cmd.append("--add_zero_vds")
        info = "  ".join(f"{k}={r.get(k)}" for k in
                         ("category", "ids_rmse", "combined_gm", "knee_combiner",
                          "knee_alpha_scale", "knee_vgs_thr") if r.get(k) not in (None, ""))
        print(f"\n[{label}] {run_dir.name}\n    {info}\n    {' '.join(cmd)}", flush=True)
        if args.dry_run:
            continue
        rc = subprocess.call(cmd)
        if rc == 0:
            ok += 1
            out_dir.mkdir(parents=True, exist_ok=True)
            for src_name in ("plot_saved_state_full.png",
                             "plot_saved_state_full_eq_comparison.png",
                             "plot_saved_state_full_val.png",
                             "plot_saved_state_full.md",
                             "plot_saved_state_full.va"):
                src = run_dir / src_name
                if src.is_file():
                    dst = out_dir / src_name
                    shutil.copy2(src, dst)
                    print(f"[{label}] copied -> {dst}", flush=True)
                    if args.open and src_name == "plot_saved_state_full.png":
                        try:
                            os.startfile(str(dst))   # open the main plot in the default viewer
                            time.sleep(0.7)          # see comment on the other os.startfile call
                        except Exception as e:
                            print(f"[{label}] could not open {dst}: {e}", flush=True)
            run_jsons = sorted(run_dir.glob("run_*.json"), key=lambda p: p.stat().st_mtime)
            if run_jsons:
                src_json = run_jsons[-1]
                shutil.copy2(src_json, out_dir / src_json.name)
                print(f"[{label}] copied -> {out_dir / src_json.name}", flush=True)
        else:
            fail += 1
            print(f"[{label}] FAILED rc={rc}", flush=True)

    if not args.dry_run:
        print(f"\ndone. ok={ok} fail={fail}")


if __name__ == "__main__":
    main()
